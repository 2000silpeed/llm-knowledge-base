"""Excel 인제스터 (W1-03)

openpyxl로 시트별 파싱 → 마크다운 테이블 변환
수식: 계산값 + [formula: ...] 주석 병기
차트: 차트 메타데이터 텍스트 설명 (Vision API 미지원 — xlsx 차트는 렌더링 이미지 없음)
청킹: 1000행 단위 분할, 컬럼 헤더 반복 포함
출력: raw/office/{파일명}.md + raw/office/{파일명}.meta.yaml

사용 예:
    from scripts.ingest_excel import ingest_excel
    result = ingest_excel("/path/to/file.xlsx")
    # {"status": "ok", "path": "raw/office/file.md", "sheets": 3, ...}
"""

import hashlib
import logging
import re
from datetime import datetime, timezone
from pathlib import Path

import yaml

from scripts.token_counter import estimate_tokens, load_settings
from scripts.utils import slugify as _slugify

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────
# 셀 값 처리
# ──────────────────────────────────────────────

def _cell_display(cell_formula, cell_value) -> str:
    """수식셀: '계산값 [formula: =수식]', 일반셀: '값' 문자열 반환합니다."""
    # 수식 여부 확인
    is_formula = isinstance(cell_formula, str) and cell_formula.startswith("=")

    if cell_value is None:
        raw_val = ""
    elif isinstance(cell_value, float):
        # 정수로 표현 가능하면 소수점 제거
        raw_val = str(int(cell_value)) if cell_value == int(cell_value) else str(cell_value)
    elif isinstance(cell_value, datetime):
        raw_val = cell_value.strftime("%Y-%m-%d %H:%M:%S")
    else:
        raw_val = str(cell_value)

    # 마크다운 테이블 파이프 이스케이프
    raw_val = raw_val.replace("|", "\\|").replace("\n", " ").replace("\r", "")

    if is_formula:
        formula_str = str(cell_formula).replace("|", "\\|")
        return f"{raw_val} `[formula: {formula_str}]`" if raw_val else f"`[formula: {formula_str}]`"

    return raw_val


# ──────────────────────────────────────────────
# 시트 → 마크다운 테이블
# ──────────────────────────────────────────────

def _sheet_to_markdown(
    ws_formula,
    ws_value,
    rows_per_chunk: int,
) -> str:
    """시트 두 뷰(수식/값)를 읽어 마크다운 섹션 문자열을 반환합니다.

    rows_per_chunk 행 단위로 분할하고, 각 청크에 컬럼 헤더를 반복합니다.
    """
    # 데이터 영역 추출
    formula_rows = list(ws_formula.iter_rows(values_only=False))
    value_rows = list(ws_value.iter_rows(values_only=True))

    if not formula_rows:
        return "_시트가 비어 있습니다._\n"

    # 실제 데이터가 있는 행/열 범위 계산
    max_col = max(
        (ws_formula.max_column or 1),
        (ws_value.max_column or 1),
    )

    # 첫 행 → 헤더
    header_cells_formula = formula_rows[0]
    header_cells_value = value_rows[0] if value_rows else []

    headers: list[str] = []
    for col_idx in range(max_col):
        f_cell = header_cells_formula[col_idx] if col_idx < len(header_cells_formula) else None
        v_val = header_cells_value[col_idx] if col_idx < len(header_cells_value) else None

        f_val = f_cell.value if f_cell is not None else None
        display = _cell_display(f_val, v_val)
        headers.append(display or f"열{col_idx + 1}")

    header_line = "| " + " | ".join(headers) + " |"
    separator_line = "| " + " | ".join(["---"] * max_col) + " |"

    # 데이터 행
    data_formula_rows = formula_rows[1:]
    data_value_rows = value_rows[1:] if len(value_rows) > 1 else []

    total_rows = len(data_formula_rows)

    if total_rows == 0:
        return f"{header_line}\n{separator_line}\n"

    # 청크 분할
    sections: list[str] = []
    chunk_count = max(1, -(-total_rows // rows_per_chunk))  # ceil division

    for chunk_idx in range(chunk_count):
        start = chunk_idx * rows_per_chunk
        end = min(start + rows_per_chunk, total_rows)
        chunk_formula = data_formula_rows[start:end]
        chunk_value = data_value_rows[start:end] if data_value_rows else []

        if chunk_count > 1:
            section_header = (
                f"\n**행 {start + 1}–{end} / 전체 {total_rows}행 (청크 {chunk_idx + 1}/{chunk_count})**\n\n"
            )
        else:
            section_header = ""

        lines: list[str] = [section_header, header_line, separator_line]

        for row_idx, formula_row in enumerate(chunk_formula):
            v_row = chunk_value[row_idx] if row_idx < len(chunk_value) else []
            cells: list[str] = []
            for col_idx in range(max_col):
                f_cell = formula_row[col_idx] if col_idx < len(formula_row) else None
                v_val = v_row[col_idx] if col_idx < len(v_row) else None

                f_val = f_cell.value if f_cell is not None else None
                cells.append(_cell_display(f_val, v_val))

            lines.append("| " + " | ".join(cells) + " |")

        sections.append("\n".join(lines))

    return "\n\n".join(sections) + "\n"


# ──────────────────────────────────────────────
# 차트 메타데이터 추출
# ──────────────────────────────────────────────

def _extract_chart_descriptions(ws) -> list[str]:
    """시트의 차트 목록에서 텍스트 설명을 생성합니다.

    xlsx 차트는 렌더링된 이미지가 없으므로 메타데이터(타입, 제목, 시리즈)로 설명합니다.
    """
    descriptions: list[str] = []
    if not hasattr(ws, "_charts"):
        return descriptions

    for i, chart in enumerate(ws._charts, start=1):
        lines: list[str] = []
        chart_type = type(chart).__name__.replace("Chart", "")

        title = ""
        try:
            if chart.title:
                if isinstance(chart.title, str):
                    title = chart.title
                else:
                    # openpyxl Title 객체
                    title = str(chart.title)
        except Exception:
            pass

        lines.append(f"**[차트 {i}]** 유형: {chart_type}" + (f" / 제목: {title}" if title else ""))

        # 시리즈 정보
        try:
            series_list = chart.series
            if series_list:
                for s_idx, series in enumerate(series_list, start=1):
                    s_title = ""
                    try:
                        if hasattr(series, "title") and series.title:
                            s_title = str(series.title)
                    except Exception:
                        pass
                    lines.append(f"  - 시리즈 {s_idx}" + (f": {s_title}" if s_title else ""))
        except Exception:
            pass

        descriptions.append("\n".join(lines))

    return descriptions


# ──────────────────────────────────────────────
# 퍼블릭 진입점
# ──────────────────────────────────────────────

def ingest_excel(
    excel_path: str | Path,
    project_root: Path | str | None = None,
    settings: dict | None = None,
) -> dict:
    """Excel 파일(.xlsx / .xls)을 인제스트합니다.

    Args:
        excel_path: Excel 파일 경로
        project_root: 프로젝트 루트 경로. None이면 이 파일 기준 상위 디렉토리.
        settings: 설정 dict. None이면 settings.yaml 자동 로드.

    Returns:
        {
            "status": "ok" | "error",
            "path": str,          # 저장된 .md 파일 경로 (프로젝트 루트 기준)
            "meta_path": str,     # 저장된 .meta.yaml 파일 경로
            "title": str,
            "sheets": int,
            "token_count": int,
            "message": str,       # 오류 시 메시지
        }
    """
    try:
        import openpyxl
    except ImportError:
        return {"status": "error", "message": "openpyxl이 설치되지 않았습니다. `uv add openpyxl`"}

    if project_root is None:
        project_root = Path(__file__).parent.parent
    project_root = Path(project_root)
    excel_path = Path(excel_path)

    if not excel_path.exists():
        return {"status": "error", "message": f"파일 없음: {excel_path}"}
    if excel_path.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
        return {"status": "error", "message": f"Excel 파일이 아닙니다: {excel_path}"}

    if settings is None:
        settings = load_settings()

    office_dir = project_root / settings["paths"]["raw"] / "office"
    office_dir.mkdir(parents=True, exist_ok=True)

    rows_per_chunk = settings.get("chunking", {}).get("excel_rows_per_chunk", 1000)

    logger.info("Excel 로드 중: %s", excel_path)

    try:
        # 수식 뷰: 수식 문자열 그대로 (data_only=False)
        wb_formula = openpyxl.load_workbook(str(excel_path), data_only=False, read_only=True)
        # 값 뷰: 마지막 저장 시 캐시된 계산값 (data_only=True)
        wb_value = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)
    except Exception as exc:
        return {"status": "error", "message": f"Excel 열기 실패: {exc}"}

    collected_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    title = excel_path.stem

    sheet_names = wb_formula.sheetnames
    sheet_count = len(sheet_names)

    all_sections: list[str] = []
    sheet_meta: list[dict] = []

    for sheet_name in sheet_names:
        ws_formula = wb_formula[sheet_name]
        ws_value = wb_value[sheet_name]

        logger.debug("시트 처리 중: %s", sheet_name)

        section_lines: list[str] = [f"## {sheet_name}\n"]

        # 차트 설명 (read_only=True에서는 _charts 접근 불가 → 별도 로드)
        chart_descriptions: list[str] = []
        try:
            import openpyxl as _opx
            wb_charts = _opx.load_workbook(str(excel_path), data_only=False, read_only=False)
            ws_charts = wb_charts[sheet_name]
            chart_descriptions = _extract_chart_descriptions(ws_charts)
            wb_charts.close()
        except Exception as exc:
            logger.debug("차트 추출 건너뜀 (%s): %s", sheet_name, exc)

        if chart_descriptions:
            section_lines.append("### 차트\n")
            section_lines.extend(desc + "\n" for desc in chart_descriptions)
            section_lines.append("")

        # 테이블 데이터
        section_lines.append("### 데이터\n")
        table_md = _sheet_to_markdown(ws_formula, ws_value, rows_per_chunk)
        section_lines.append(table_md)

        section_text = "\n".join(section_lines)
        all_sections.append(section_text)

        sheet_meta.append({
            "name": sheet_name,
            "rows": (ws_formula.max_row or 0) - 1,  # 헤더 제외
            "cols": ws_formula.max_column or 0,
            "charts": len(chart_descriptions),
        })

    wb_formula.close()
    wb_value.close()

    body = "\n\n---\n\n".join(all_sections)
    token_count = estimate_tokens(body)

    frontmatter: dict = {
        "title": title,
        "source_file": str(excel_path),
        "collected_at": collected_at,
        "sheets": sheet_count,
        "token_count": token_count,
    }
    fm_str = yaml.dump(frontmatter, allow_unicode=True, default_flow_style=False, sort_keys=False)
    document = f"---\n{fm_str}---\n\n# {title}\n\n{body}\n"

    # 파일명 결정 (충돌 방지)
    stem = _slugify(excel_path.stem) or "spreadsheet"
    filename = f"{stem}.md"
    dest = office_dir / filename
    if dest.exists():
        file_hash = hashlib.md5(excel_path.stem.encode()).hexdigest()[:6]
        filename = f"{stem}_{file_hash}.md"
        dest = office_dir / filename

    dest.write_text(document, encoding="utf-8")
    rel_path = str(dest.relative_to(project_root))

    # .meta.yaml 저장
    meta: dict = {
        "source_file": str(excel_path),
        "collected_at": collected_at,
        "title": title,
        "sheets": sheet_meta,
        "token_count": token_count,
        "rows_per_chunk": rows_per_chunk,
    }
    meta_filename = dest.stem + ".meta.yaml"
    meta_dest = office_dir / meta_filename
    meta_dest.write_text(
        yaml.dump(meta, allow_unicode=True, default_flow_style=False, sort_keys=False),
        encoding="utf-8",
    )
    meta_rel_path = str(meta_dest.relative_to(project_root))

    logger.info(
        "저장 완료: %s (시트: %d, 토큰: %d)", rel_path, sheet_count, token_count
    )

    return {
        "status": "ok",
        "path": rel_path,
        "meta_path": meta_rel_path,
        "title": title,
        "sheets": sheet_count,
        "token_count": token_count,
    }
